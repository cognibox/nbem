import argparse
import csv
import re
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from fuzzywuzzy import fuzz
from datetime import datetime

CBX_HEADERS_LENGTH = 7
CBX_FIRSTNAME, CBX_LASTNAME, CBX_ID, CBX_BIRTHDATE, CBX_COMPANY, CBX_PARENTS, CBX_PREVIOUS = range(CBX_HEADERS_LENGTH)

cbx_headers = ['first_name', 'last_name', 'cbx_id', 'birth_date', 'contractor', 'parents', 'previous']

HC_HEADERS_LENGTH = 3
HC_COMPANY, HC_FIRSTNAME, HC_LASTNAME = range(HC_HEADERS_LENGTH)

hc_headers = ['company', 'first_name', 'last_name']

# noinspection SpellCheckingInspection
BASE_GENERIC_COMPANY_NAME_WORDS = ['construction', 'contracting', 'industriel', 'industriels', 'service',
                                   'services', 'inc', 'limited', 'ltd', 'ltee', 'ltée', 'co', 'industrial',
                                   'solutions', 'llc', 'enterprises', 'systems', 'industries',
                                   'technologies', 'company', 'corporation', 'installations', 'enr']


analysis_headers = ['cbx_id', 'birthdate', 'best_score', 'partial', 'same_birthdate', 'match_count', 'analysis']


def chunks(lst, n):
    """Yield successive n-sized chunks from lst."""
    for i in range(0, len(lst), n):
        yield lst[i:i + n]


cbx_headers_text = '\n'.join([', '.join(x) for x in list(chunks(cbx_headers, 5))])
hc_headers_text = '\n'.join([', '.join(x) for x in list(chunks(hc_headers, 5))])
analysis_headers_text = '\n'.join([', '.join(x) for x in list(chunks(analysis_headers, 5))])


# define commandline parser
parser = argparse.ArgumentParser(description='Tool to match employees without birthday to employees ID in CBX, '
                                             'all input/output files must be in the current directory',
                                 formatter_class=argparse.RawTextHelpFormatter)
parser.add_argument('cbx_list',
                    help=f'csv DB export file (no header) of employees with the following '
                         f'columns:\n{cbx_headers_text}\n\n')


parser.add_argument('hc_list',
                    help=f'csv file (with header) and the following columns:\n{hc_headers_text}\n\n'
                         'Followed by any other columns...'
                    )
# noinspection SpellCheckingInspection
parser.add_argument('output',
                    help=f'csv file with the hc_list columns and following columns\n{analysis_headers_text}\n\n' 
                         '''Matching information format:
    Cognibox ID, firstname lastname, birthdate --> Contractor 1
    [parents: C1 parent1;C1 parent2;etc..] 
    [previous: Empl. Previous1;Empl. Previous2], match ratio 1,
    Contractor 2 [C2 parent1;C2 parent2;etc..], match ratio 2, etc...
    The matching ratio is a value betwween 0 and 100, where 100 is a perfect match.
Please note the Cognibox ID and birthdate is set ONLY if a single match his found. If no match
or multiple matches are found it is left empty.''')

parser.add_argument('--cbx_list_encoding', dest='cbx_encoding', action='store',
                    default='utf-8',
                    help='Encoding for the cbx list (default: utf-8)')

parser.add_argument('--min_company_match_ratio', dest='ratio_company', action='store',
                    default=60,
                    help='Minimum match ratio for contractors, between 0 and 100 (default 60)')

parser.add_argument('--list_separator', dest='list_separator', action='store',
                    default=';',
                    help='string separator used for lists (default: ;)')

parser.add_argument('--additional_generic_name_word', dest='additional_generic_name_word', action='store',
                    default='',
                    help='list of generic words in company name to ignore separated by the list separator'
                         ' (default separator is ;)')

# parser.add_argument('--min_name_match_ratio', dest='ratio_name', action='store',
#                    default=90,
#                    help='Minimum match ratio for contractors, between 0 and 100 (default 90)')

parser.add_argument('--min_first_name_match_ratio', dest='ratio_first_name', action='store',
                    default=80,
                    help='Minimum match ratio for first name, between 0 and 100 (default 80)')

parser.add_argument('--min_last_name_match_ratio', dest='ratio_last_name', action='store',
                    default=90,
                    help='Minimum match ratio for last name, between 0 and 100 (default 90)')

parser.add_argument('--no_headers', dest='no_headers', action='store_true',
                    help='to indicate that input files have no headers')

parser.add_argument('--ignore_warnings', dest='ignore_warnings', action='store_true',
                    help='to ignore data consistency checks and run anyway...')

args = parser.parse_args()

if len(hc_headers) != HC_HEADERS_LENGTH:
    raise AssertionError('hc header inconsistencies')

if len(cbx_headers) != CBX_HEADERS_LENGTH:
    raise AssertionError('cbx header inconsistencies')

GENERIC_COMPANY_NAME_WORDS = BASE_GENERIC_COMPANY_NAME_WORDS + \
                             args.additional_generic_name_word.split(args.list_separator)


def remove_generics(company_name):
    for word in GENERIC_COMPANY_NAME_WORDS:
        company_name = re.sub(r'\b' + word + r'\b', '', company_name)
    return company_name


# noinspection PyShadowingNames
def check_headers(headers, standards, ignore):
    headers = [x.lower().strip() for x in headers]
    for idx, val in enumerate(standards):
        if val != headers[idx]:
            print(f'WARNING: got "{headers[idx]}" while expecting "{val}" in column {idx + 1}')
            if not ignore:
                exit(-1)


def clean_company_name(name):
    name = name.lower().replace('.', '').replace(',', '').strip()
    name = re.sub(r"\([^()]*\)", "", name)
    name = remove_generics(name)
    return name


if __name__ == '__main__':
    data_path = './data/'
    cbx_file = data_path + args.cbx_list
    hc_file = data_path + args.hc_list
    output_file = data_path + args.output

    # output parameters used
    print(f'Starting at {datetime.now()}')
    print(f'Reading CBX list: {args.cbx_list} [{args.cbx_encoding}]')
    print(f'Reading HC list: {args.hc_list}')
    print(f'Outputting results in: {args.output}')
    print(f'contractor match ratio: {args.ratio_company}')
    print(f'employee first name match ratio: {args.ratio_first_name}')
    print(f'employee last name match ratio: {args.ratio_last_name}')
    # read data
    cbx_data = []
    hc_data = []
    print('Reading Cognibox data file...')
    with open(cbx_file, 'r', encoding=args.cbx_encoding) as cbx:
        for row in csv.reader(cbx):
            cbx_data.append(row)
    print(f'Completed reading {len(cbx_data)} employees.')
    print('Reading hiring client data file...')
    hc_wb = openpyxl.load_workbook(hc_file, read_only=True)
    hc_sheet = hc_wb.active
    max_row = hc_sheet.max_row
    max_column = hc_sheet.max_column
    if max_column > 250 or max_row > 10000:
        print(f'WARNING: File is large: {max_row} rows and {max_column}. must be less than 10000 and 250')
        if not args.ignore_warnings:
            exit(-1)
    for row in hc_sheet.rows:
        if not row[0].value:
            continue
        hc_data.append([cell.value if cell.value else '' for cell in row])
    print(f'Completed reading {len(hc_data)} employees.')

    print(f'Starting analysis...')
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "results"

    if not args.no_headers:
        headers = hc_data.pop(0)
        check_headers(headers, hc_headers, args.ignore_warnings)
        headers.extend(analysis_headers)
        for index, value in enumerate(headers):
            out_ws.cell(1, index+1, value)
        out_wb.save(filename=output_file)
        headers = cbx_data.pop(0)
        check_headers(headers, cbx_headers, args.ignore_warnings)

    # match
    total = len(hc_data)
    index = 1
    hc_row = []
    for hc_row in hc_data:
        matches = []
        hc_firstname = str(hc_row[HC_FIRSTNAME]).lower().strip()
        hc_lastname = str(hc_row[HC_LASTNAME]).lower().strip()
        hc_company = str(hc_row[HC_COMPANY])
        hc_company_cleaned = clean_company_name(hc_company)

        for cbx_row in cbx_data:
            cbx_firstname = str(cbx_row[CBX_FIRSTNAME]).lower().strip()
            cbx_lastname = str(cbx_row[CBX_LASTNAME]).lower().strip()
            if not hc_firstname:
                cbx_lastname = ' '.join([cbx_firstname, cbx_lastname])
                cbx_firstname = ''
            cbx_company = str(cbx_row[CBX_COMPANY])
            cbx_company_cleaned = clean_company_name(cbx_company)

            cbx_parents = cbx_row[CBX_PARENTS]
            cbx_previous = cbx_row[CBX_PREVIOUS]
            ratio_first_name = fuzz.token_set_ratio(cbx_firstname, hc_firstname)
            ratio_last_name = fuzz.token_sort_ratio(cbx_lastname, hc_lastname)
            if ratio_first_name >= float(args.ratio_first_name) and ratio_last_name >= float(args.ratio_last_name):
                ratio_first_name_exact = fuzz.token_sort_ratio(cbx_firstname, hc_firstname)
                partial = True if ratio_first_name_exact < float(args.ratio_first_name) else False
                ratio_company = fuzz.token_sort_ratio(cbx_company_cleaned, hc_company_cleaned)
                ratio_parent = 0
                cbx_parent_list = cbx_parents.split(args.list_separator)
                for item in cbx_parent_list:
                    if item == cbx_company:
                        continue
                    ratio = fuzz.token_sort_ratio(clean_company_name(item), hc_company_cleaned)
                    ratio_parent = ratio if ratio > ratio_parent else ratio_parent
                ratio_previous = 0
                for item in cbx_previous.split(args.list_separator):
                    if item == cbx_company or item in cbx_parent_list:
                        continue
                    ratio = fuzz.token_sort_ratio(clean_company_name(item), hc_company_cleaned)
                    ratio_previous = ratio if ratio > ratio_previous else ratio_previous
                if (ratio_company >= float(args.ratio_company) or
                        ratio_parent >= float(args.ratio_company) or
                        ratio_previous >= float(args.ratio_company)):
                    print('   --> ', cbx_firstname, cbx_lastname, cbx_company, cbx_row[CBX_ID], ratio_company,
                          ratio_parent, ratio_previous)
                    ratio_company = ratio_parent if ratio_parent > ratio_company else ratio_company
                    ratio_company = ratio_previous if ratio_previous > ratio_company else ratio_company
                    overall_ratio = ratio_company * ratio_first_name * ratio_last_name / 10000
                    parent_str = f'[parent: {cbx_parents}]' if cbx_parents else None
                    previous_str = f'[previous: {cbx_previous}]' if cbx_previous else None
                    display = [cbx_company]
                    if parent_str:
                        display.append(parent_str)
                    if previous_str:
                        display.append(previous_str)
                    cbx_company = ' '.join(display)
                    matches.append({'cbx_id': cbx_row[CBX_ID],
                                    'firstname': cbx_firstname,
                                    'lastname': cbx_lastname,
                                    'birthdate': cbx_row[CBX_BIRTHDATE],
                                    'company': cbx_company,
                                    'ratio': overall_ratio,
                                    'partial': partial, })
        ids = []
        best_match = 0
        matches.sort(key=lambda x: x['ratio'], reverse=True)
        companies = []
        if matches:
            best_match = matches[0]['ratio'] if matches[0]['ratio'] > best_match else best_match
        for item in matches[0:5]:
            companies.append(f'{item["company"]}: {item["ratio"]}')
            ids.append(f'{item["cbx_id"]}, {item["firstname"]} {item["lastname"]},'
                       f'{item["birthdate"]} --> {", ".join(companies)}')

        uniques_cbx_id = set(item['cbx_id'] for item in matches)
        same_bd = ''
        if len(uniques_cbx_id) >= 1:
            same_bd = True
            if len(uniques_cbx_id) != 1:
                bd = matches[0]['birthdate']
                for item in matches[1:]:
                    if item['birthdate'] != bd:
                        same_bd = False
                        break
        # append matching results to the hc_list
        hc_row.append(matches[0]["cbx_id"] if len(uniques_cbx_id) == 1 else '?' if len(uniques_cbx_id) > 1 else '')
        hc_row.append(matches[0]["birthdate"] if len(uniques_cbx_id) == 1 else '')
        hc_row.append(best_match if len(uniques_cbx_id) == 1 else '')
        hc_row.append(matches[0]['partial'] if len(uniques_cbx_id) == 1 else '')
        hc_row.append(same_bd)
        hc_row.append(len(uniques_cbx_id) if len(uniques_cbx_id) else '')
        hc_row.append('\n'.join(ids))
        for i, value in enumerate(hc_row):
            out_ws.cell(index+1, i+1, value)
        if index % 10:
            out_wb.save(filename=output_file)
        print(f'{index} of {total} [{len(uniques_cbx_id)} found]')
        index += 1

    # formatting the excel...
    tab = Table(displayName='results', ref=f'A1:{get_column_letter(len(hc_row))}{len(hc_data)+1}')
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    dims = {}
    for row in out_ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        out_ws.column_dimensions[col].width = value
    out_ws.column_dimensions[get_column_letter(HC_HEADERS_LENGTH+len(analysis_headers)-6)].width = 150
    for i in range(2, len(hc_data)+1):
        out_ws.cell(i, HC_HEADERS_LENGTH+len(analysis_headers)-6).alignment = Alignment(wrapText=True)
    tab.tableStyleInfo = style
    out_ws.add_table(tab)
    out_wb.save(filename=output_file)
    print('Analysis Completed')
    print(f'Completed at {datetime.now()}')
